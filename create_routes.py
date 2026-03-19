from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Routes"

# Headers
headers = ["id", "number", "departure", "destination", "airport_name", "flight_time", "notes", "section"]
sheet.append(headers)

# Format headers
header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Flight routes data
routes = [
  {"id": 1, "number": 1, "departure": "LLBG", "destination": "LCEN", "airport_name": "לרנקה, קפריסין", "flight_time": 65, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 2, "number": 2, "departure": "LCEN", "destination": "LGSR", "airport_name": "רודוס, יוון", "flight_time": 60, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 3, "number": 3, "departure": "LGSR", "destination": "LGKF", "airport_name": "קורפו, יוון", "flight_time": 100, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 4, "number": 4, "departure": "LGKF", "destination": "LIRF", "airport_name": "רומא, איטליה", "flight_time": 95, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 5, "number": 5, "departure": "LIRF", "destination": "LFMN", "airport_name": "ניס, צרפת", "flight_time": 70, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 6, "number": 6, "departure": "LFMN", "destination": "LFPB", "airport_name": "פריז לה בורז'ה", "flight_time": 95, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 7, "number": 7, "departure": "LFPB", "destination": "EGBB", "airport_name": "בירמינגהם, בריטניה", "flight_time": 80, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 8, "number": 8, "departure": "EGBB", "destination": "EIDW", "airport_name": "דבלין, אירלנד", "flight_time": 70, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 9, "number": 9, "departure": "EIDW", "destination": "EKVG", "airport_name": "ווגאר, איי פארו", "flight_time": 100, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 10, "number": 10, "departure": "EKVG", "destination": "BIKF", "airport_name": "קפלאוויק, איסלנד", "flight_time": 105, "notes": "", "section": "מקטע א': אירופה"},
  {"id": 11, "number": 11, "departure": "BIKF", "destination": "CYYT", "airport_name": "סנט ג'ונס, ניופאונדלנד", "flight_time": 255, "notes": "חריגה 1: 120 דקות (מעבר מגבלת 135)", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 12, "number": 12, "departure": "CYYT", "destination": "CYQX", "airport_name": "גנדר, קנדה", "flight_time": 60, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 13, "number": 13, "departure": "CYQX", "destination": "CYGZ", "airport_name": "גרינלנד, קנדה", "flight_time": 100, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 14, "number": 14, "departure": "CYGZ", "destination": "CYQB", "airport_name": "קוויבק סיטי, קנדה", "flight_time": 85, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 15, "number": 15, "departure": "CYQB", "destination": "CYYZ", "airport_name": "טורונטו, קנדה", "flight_time": 90, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 16, "number": 16, "departure": "CYYZ", "destination": "KDTW", "airport_name": "דטרויט, ארה\"ב", "flight_time": 65, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 17, "number": 17, "departure": "KDTW", "destination": "KORD", "airport_name": "שיקגו, ארה\"ב", "flight_time": 80, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 18, "number": 18, "departure": "KORD", "destination": "KOMA", "airport_name": "אומהה, ארה\"ב", "flight_time": 95, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 19, "number": 19, "departure": "KOMA", "destination": "KDEN", "airport_name": "דנבר, ארה\"ב", "flight_time": 105, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 20, "number": 20, "departure": "KDEN", "destination": "KSLC", "airport_name": "סולט לייק סיטי, ארה\"ב", "flight_time": 85, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 21, "number": 21, "departure": "KSLC", "destination": "KLAS", "airport_name": "לאס וגאס, ארה\"ב", "flight_time": 80, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 22, "number": 22, "departure": "KLAS", "destination": "KLAX", "airport_name": "לוס אנג'לס, ארה\"ב", "flight_time": 60, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 23, "number": 23, "departure": "KLAX", "destination": "KSAN", "airport_name": "סן דייגו, ארה\"ב", "flight_time": 60, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 24, "number": 24, "departure": "KSAN", "destination": "KPHX", "airport_name": "פיניקס, ארה\"ב", "flight_time": 85, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 25, "number": 25, "departure": "KPHX", "destination": "KDFW", "airport_name": "דאלאס, ארה\"ב", "flight_time": 130, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 26, "number": 26, "departure": "KDFW", "destination": "KHOU", "airport_name": "יוסטון, ארה\"ב", "flight_time": 75, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 27, "number": 27, "departure": "KHOU", "destination": "KMIA", "airport_name": "מיאמי, ארה\"ב", "flight_time": 120, "notes": "", "section": "מקטע ב': חצייה אטלנטית וצפון אמריקה"},
  {"id": 28, "number": 28, "departure": "KMIA", "destination": "MMMX", "airport_name": "מקסיקו סיטי, מקסיקו", "flight_time": 120, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 29, "number": 29, "departure": "MMMX", "destination": "MRPV", "airport_name": "מוראלס, גוואטמלה", "flight_time": 100, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 30, "number": 30, "departure": "MRPV", "destination": "MDSSD", "airport_name": "סן סלבדור, אל סלבדור", "flight_time": 60, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 31, "number": 31, "departure": "MDSSD", "destination": "MHTG", "airport_name": "טגוסיגאלפה, הונדורס", "flight_time": 75, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 32, "number": 32, "departure": "MHTG", "destination": "MNMG", "airport_name": "מנגווה, ניקרגווה", "flight_time": 80, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 33, "number": 33, "departure": "MNMG", "destination": "MPMG", "airport_name": "פנמה סיטי, פנמה", "flight_time": 90, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 34, "number": 34, "departure": "MPMG", "destination": "SKBO", "airport_name": "בוגוטה, קולומביה", "flight_time": 95, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 35, "number": 35, "departure": "SKBO", "destination": "SEQU", "airport_name": "קיטו, אקוודור", "flight_time": 85, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 36, "number": 36, "departure": "SEQU", "destination": "SPIM", "airport_name": "לימה, פרו", "flight_time": 100, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 37, "number": 37, "departure": "SPIM", "destination": "SLET", "airport_name": "לה פאז, בוליביה", "flight_time": 95, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 38, "number": 38, "departure": "SLET", "destination": "SABP", "airport_name": "בואנוס איירס, ארגנטינה", "flight_time": 150, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 39, "number": 39, "departure": "SABP", "destination": "SCCI", "airport_name": "קונספסיון, צ'ילה", "flight_time": 120, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 40, "number": 40, "departure": "SCCI", "destination": "SCPQ", "airport_name": "פונטה ארנס, צ'ילה", "flight_time": 80, "notes": "", "section": "מקטע ג': אמריקה הדרומית"},
  {"id": 41, "number": 41, "departure": "SCPQ", "destination": "NZCH", "airport_name": "כריסטצ'רץ', ניו זילנד", "flight_time": 310, "notes": "חריגה 2: עברת אנטארקטיקה", "section": "מקטע ד': אנטארקטיקה וסיידני"},
  {"id": 42, "number": 42, "departure": "NZCH", "destination": "YSSY", "airport_name": "סידני, אוסטרליה", "flight_time": 165, "notes": "", "section": "מקטע ד': אנטארקטיקה וסיידני"},
  {"id": 43, "number": 43, "departure": "YSSY", "destination": "RPLL", "airport_name": "מנילה, פיליפינים", "flight_time": 200, "notes": "", "section": "מקטע ה': אסיה"},
  {"id": 44, "number": 44, "departure": "RPLL", "destination": "ZSSS", "airport_name": "שנגחאי, סין", "flight_time": 160, "notes": "", "section": "מקטע ה': אסיה"},
  {"id": 45, "number": 45, "departure": "ZSSS", "destination": "RJTT", "airport_name": "טוקיו, יפן", "flight_time": 125, "notes": "", "section": "מקטע ה': אסיה"},
  {"id": 46, "number": 46, "departure": "RJTT", "destination": "VHHH", "airport_name": "הונג קונג, הונג קונג", "flight_time": 170, "notes": "", "section": "מקטע ה': אסיה"},
  {"id": 47, "number": 47, "departure": "VHHH", "destination": "LLBG", "airport_name": "לרנקה/תל אביב, ישראל", "flight_time": 280, "notes": "חזרה הביתה!", "section": "חזרה הביתה"}
]

# Add data rows
for route in routes:
    row_data = [
        route["id"],
        route["number"],
        route["departure"],
        route["destination"],
        route["airport_name"],
        route["flight_time"],
        route["notes"],
        route["section"]
    ]
    sheet.append(row_data)

# Format data rows
for row_num in range(2, len(routes) + 2):
    for col_num in range(1, len(headers) + 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.border = border
        cell.alignment = Alignment(horizontal='right' if col_num == 5 else 'center', vertical='center')

# Set column widths
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 8
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 25
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 30
sheet.column_dimensions['H'].width = 30

# Save
wb.save('C:/cluade/world travel/Doctor_Simulator_Routes.xlsx')
print("Excel file created successfully!")
